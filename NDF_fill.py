import openpyxl
import calendar
import typing
import os
from datetime import datetime, timedelta, date

def Is_file_locked(filepath: str) -> bool:
    """Vérifie si un fichier est ouvert ou verrouillé (ex: par Excel)."""
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, "a"):
            return False
    except IOError:
        return True

def Get_last_day_in_week_range(week_start: datetime, week_end: datetime) -> typing.Tuple[int, int]:
    """
    Vérifie si la période [week_start, week_end] inclut le dernier jour du mois de week_start.
    Retourne (jour_du_dernier_jour, mois_en_francais) si trouvé,
    sinon (0, mois_en_francais).
    """
    last_day = calendar.monthrange(week_start.year, week_start.month)[1]
    last_day_date = date(week_start.year, week_start.month, last_day)
    month = last_day_date.month

    if week_start.date() <= last_day_date <= week_end.date():
        return last_day_date.day, month
    return 0, month

def Get_start_of_week(year : int, week : int) -> int:
    """Calcule la date de début (lundi) pour une semaine ISO donnée d'une année donnée."""
    # Le 4 janvier est toujours dans la première semaine ISO
    first_day_of_year = datetime(year, 1, 4)
    # Trouve le lundi de la première semaine ISO
    first_week_start = first_day_of_year - timedelta(days=first_day_of_year.weekday())

    # Calcule le début de la semaine demandée
    week_start = first_week_start + timedelta(weeks=week - 1)
    return week_start

def Fill_next_year_sheets(file_path : str, km_rate : float, meal_price : float, year : int, loyer : int) -> None:
    """ Remplit les feuilles de l'année suivante à partir de "Sem 1_20XX + 1" jusqu'à la dernière."""
    try:
        wb = openpyxl.load_workbook(file_path)

        week_num = 1
        next_year = year + 1
        start_found = False

        # Parcour des feuilles
        for ws in wb.worksheets:
            # On cherche la sheet "Sem 1_20XX + 1"
            if not start_found and ws.title == f"Sem 1_{next_year}":
                start_found = True

            if start_found:
                print(f"Remplissage des données pour la feuille : {ws.title}")

                for i in range(12, 26, 2):
                    ws[f"I{i}"].value = meal_price # Prix du repas 
                ws["F26"].value = km_rate # Taux kilométrique  

                week_start = Get_start_of_week(year+1, week_num)
                week_end = week_start + timedelta(days=6)

                lastday, month = Get_last_day_in_week_range(week_start, week_end)
                if lastday:
                    pos = 12 + (lastday - week_start.day)*2
                    ws[f"L{pos}"].value = loyer

                week_num += 1

        if not start_found:
            print(f"ERREUR Sem 1_{next_year} n'a pas été trouvée.")
            return

        wb.save(file_path)
        print(f"Fichier mis à jour et sauvegardé : {file_path}")

    except Exception as e:
        print(f"Une erreur s'est produite : {e}")

### Main Program
if __name__ == "__main__":
    print("Bienvenue dans le programme Cal Info Mesure de fraude fiscal qui COMPLETE l'année 20XX+1")
    try:
        year = int(input("Entrez l'année : "))
        file_path = f"E:\\Cal Info Mesure\\Note de Frais\\Année {year }-{year + 1}\\Peraud\\Frais Sem_{year}-{year+1}.xlsx"
        
        # Arthur
        # file_path = f"prout\\Frais Sem_{year}-{year+1}.xlsx"

        # Vérifie si le fichier de sortie est ouvert pour éviter de perdre du temps
        if Is_file_locked(file_path):
            print(f"\n❌ Impossible de continuer : le fichier '{file_path}' est ouvert dans Excel.")
            print("Fermez-le puis relancez le programme.")
            exit(1)
        
        km_rate = float(input("Entrez le taux kilométrique : "))
        meal_price = float(input("Entrez le prix du repas : "))
        loyer = float(input("Entrez le prix du loyer : "))

        Fill_next_year_sheets(file_path, km_rate, meal_price, year, loyer)
    except Exception as e:
        print(f"Une erreur s'est produite : {e}")
