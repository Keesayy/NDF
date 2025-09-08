# By Arthur Péraud
import os
import calendar
import openpyxl
import typing
from datetime import datetime, timedelta, date
from openpyxl.workbook import Workbook

MONTH_FR = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
]

def Is_file_locked(filepath: str) -> bool:
    """Vérifie si un fichier est ouvert ou verrouillé (ex: par Excel)."""
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, "a"):
            return False
    except IOError:
        return True

def Get_unique_filename(path : str) -> str:
    """Génère un nom de fichier unique en ajoutant (1), (2), etc. s'il existe déjà."""
    base, ext = os.path.splitext(path)
    i = 1
    new_path = path
    while os.path.exists(new_path):
        new_path = f"{base}({i}){ext}"
        i += 1
    return new_path

def Save_workbook_safely(wb: Workbook, output_file: str) -> None:
    """Sauvegarde fichier Excel, confirmation si existence sinon préfixe est ajouté (Get_unique_filename).
    Crée le dossier si le chemin n'existe pas.
    """

    # Vérifie et crée le dossier si nécessaire
    folder = os.path.dirname(output_file)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
        print(f"📂 Dossier créé : {folder}")

    if os.path.exists(output_file):
        confirm = input(f"\n⚠️  Le fichier '{output_file}' existe déjà. Voulez-vous l’écraser ? (o/n) : ").strip().lower()
        
        if confirm in ['o', 'y']:
            wb.save(output_file)
            print(f"✅ Fichier écrasé et sauvegardé sous {output_file}")
        elif confirm == 'n':
            new_file = Get_unique_filename(output_file)
            wb.save(new_file)
            print(f"📁 Fichier sauvegardé sous un nouveau nom : {new_file}")
        else:
            print("Réponse non reconnue, fichier non sauvegardé.")
    else:
        wb.save(output_file)
        print(f"\n✅ Fichier sauvegardé sous {output_file}")

def Get_number_of_weeks(year : int) -> int:
    """Calcule le nombre de semaines ISO pour une année donnée (52 ou 53 semaines)."""
    # Vérifie le numéro de semaine du dernier jour de l'année
    last_day_of_year = datetime(year, 12, 31)
    last_week = last_day_of_year.isocalendar()[1]

    # S'il y a 53 semaines ISO dans l'année, sinon il y en a 52
    return 53 if last_week == 53 else 52

def Get_start_of_week(year : int, week : int) -> int:
    """Calcule la date de début (lundi) pour une semaine ISO donnée d'une année donnée."""
    # Le 4 janvier est toujours dans la première semaine ISO
    first_day_of_year = datetime(year, 1, 4)
    # Trouve le lundi de la première semaine ISO
    first_week_start = first_day_of_year - timedelta(days=first_day_of_year.weekday())

    # Calcule le début de la semaine demandée
    week_start = first_week_start + timedelta(weeks=week - 1)
    return week_start

def Get_last_day_in_week_range(week_start: datetime, week_end: datetime) -> typing.Tuple[int, int]:
    """
    Vérifie si la période [week_start, week_end] inclut le dernier jour du mois de week_start.
    Retourne (jour_du_dernier_jour, mois_en_francais) si trouvé,
    sinon (0, mois_en_francais).
    """
    last_day = calendar.monthrange(week_start.year, week_start.month)[1]
    last_day_date = date(week_start.year, week_start.month, last_day)
    month = last_day_date.month

    if week_start.date() <= last_day_date <= week_end.date():
        return last_day_date.day, month
    return 0, month

def Last_week_contains_4_days_of_month(year : int, month : int = 4) -> (bool, int):
    """Retourne un tuple (True/False, numéro de la semaine) si la dernière semaine du mois contient au moins 4 jours du mois."""
    
    # Gérer le cas où le mois est décembre (12)
    if month == 12:
        next_month = 1
        year += 1  # Passage à l'année suivante
    else:
        next_month = month + 1
    
    # Dernier jour du mois
    last_day_of_month = datetime(year, next_month, 1) - timedelta(days=1)
    
    # Le premier jour de la semaine (lundi) pour le dernier jour du mois
    last_week_start = last_day_of_month - timedelta(days=last_day_of_month.weekday())
    
    # Calcul du nombre de jours du mois dans la dernière semaine
    days_in_last_week = 0
    for i in range(7):  
        current_day = last_week_start + timedelta(days=i)
        if current_day.month == month:
            days_in_last_week += 1
    
    # Le numéro de la semaine ISO de la dernière semaine du mois
    last_week_number = last_week_start.isocalendar()[1]
    
    # Si la dernière semaine contient 4 jours ou plus du mois
    contains_4_days = days_in_last_week >= 4
    
    return contains_4_days, last_week_number


def Create_weekly_sheets(input_file : str, year : int, km_rate : float, meal_price : float, loyer : float) -> Workbook:
    wb = openpyxl.load_workbook(input_file)
    ws_template = wb.active  # Feuille modèle par défaut

    num_weeks = Get_number_of_weeks(year)
    flag, start_week = Last_week_contains_4_days_of_month(year)
    if flag : start_week += 1
    flag, end_week_next_year = Last_week_contains_4_days_of_month(year+1)
    print("")

    # 20XX
    for week_num in range(start_week, num_weeks + 1):
        # Calcule les dates de début (lundi) et de fin (dimanche) pour chaque semaine
        week_start = Get_start_of_week(year, week_num)
        week_end = week_start + timedelta(days=6)

        # Crée une nouvelle feuille pour chaque semaine
        ws = wb.copy_worksheet(ws_template)
        ws.title = f"Sem {week_num}_{year}"

        print(f"Création de la feuille pour la semaine {week_num}, du {week_start.strftime('%d/%m/%Y')} au {week_end.strftime('%d/%m/%Y')}.")

        ws["K1"].value = week_num  # Numéro de la semaine
        ws["O5"].value = week_start.strftime("%d/%m/%Y")  # Date de début
        ws["O6"].value = week_end.strftime("%d/%m/%Y")    # Date de fin

        # Affichage Total et Mois
        lastday, month = Get_last_day_in_week_range(week_start, week_end)
        if lastday:
            flag2, _ = Last_week_contains_4_days_of_month(year, month)
            if not flag2: month = (month + 1) % 12

            # Total Péage fin du mois
            pos = 12 + (lastday - week_start.day)*2
            ws[f"E{pos}"].value = "Pe"
            ws[f"F{pos}"].font = openpyxl.styles.Font(color="FF0000", size = 8)            
            ws[f"F{pos+1}"].value = "Total Mois"
            # Loyer
            ws[f"L{pos+1}"].value = "Loyer_ES"
            ws[f"L{pos}"].value = loyer

        ws["O3"].value = MONTH_FR[month-1]
        ws["O3"].font = openpyxl.styles.Font(bold=True, color="008000", size = 9) # Texte en Gras et Vert

        k = 0
        for i in range(12, 26, 2):
            ws[f"B{i}"].value = (week_start + timedelta(days=(k))).day  # Dates
            k = k + 1

        for i in range(12, 26, 2):
            ws[f"I{i}"].value = meal_price  # Prix du repas 
        ws["F26"].value = km_rate  # Taux kilométrique     

    # 20XX +1
    if not flag : end_week_next_year -= 1
    for week_num in range(1, end_week_next_year + 1):
        # Calcule les dates de début (lundi) et de fin (dimanche) pour chaque semaine
        week_start = Get_start_of_week(year+1, week_num)
        week_end = week_start + timedelta(days=6)

        # Crée une nouvelle feuille pour chaque semaine
        ws = wb.copy_worksheet(ws_template)
        ws.title = f"Sem {week_num}_{year + 1}"

        print(f"Création de la feuille pour la semaine {week_num}, du {week_start.strftime('%d/%m/%Y')} au {week_end.strftime('%d/%m/%Y')}.")

        ws["K1"].value = week_num  # Numéro de la semaine
        ws["O5"].value = week_start.strftime("%d/%m/%Y")  # Date de début
        ws["O6"].value = week_end.strftime("%d/%m/%Y")    # Date de fin

        # Affichage Total et Mois
        lastday, month = Get_last_day_in_week_range(week_start, week_end)
        if lastday:
            flag2, _ = Last_week_contains_4_days_of_month(year + 1, month)
            if not flag2: month = (month + 1) % 12

            # Total Péage fin du mois
            pos = 12 + (lastday - week_start.day)*2
            ws[f"E{pos}"].value = "Pe"
            ws[f"F{pos}"].font = openpyxl.styles.Font(color="FF0000", size = 8)
            ws[f"F{pos+1}"].value = "Total Mois"
            # Loyer
            ws[f"L{pos+1}"].value = "Loyer_ES"
            ws[f"L{pos}"].value = None
         
        ws["O3"].value = MONTH_FR[month-1]
        ws["O3"].font = openpyxl.styles.Font(bold=True, color="008000", size = 9) # Texte en Gras et Vert

        k = 0
        for i in range(12, 26, 2):
            ws[f"B{i}"].value = (week_start + timedelta(days=(k))).day  # Dates
            k = k + 1

        ### On ne remplit pas pour l'année suivante, on modifiera plus tard les valeurs, mises à NONE    
        for i in range(12, 26, 2):
            ws[f"I{i}"].value = None  # Prix du repas 
        ws["F26"].value = None  # Taux kilométrique   

    wb.remove(ws_template)
    print(f"{num_weeks - start_week + 1 + end_week_next_year} feuilles de semaine créées.")

    return wb

### Main Program
if __name__ == "__main__":
    print("Bienvenue dans le programme Cal Info Mesure de fraude fiscal.")
    
    try:
        year = int(input("Entrez l'année du 1er Mai 20XX : "))

        input_file = "Frais Sem_Modele.xlsx"  # Fichier modèle pour le sheet modèle
        output_file = f"E:\\Cal Info Mesure\\Note de Frais\\Année {year }-{year + 1}\\Peraud\\Frais Sem_{year}-{year+1}.xlsx"

        # Arthur
        # input_file = f"Excel\\Frais Sem1_2025.xlsx"  # Fichier modèle pour le sheet modèle
        # output_file = f"prout\\Frais Sem_{year}-{year+1}.xlsx"

        # Vérifie si le fichier de sortie est ouvert pour éviter de perdre du temps
        if Is_file_locked(output_file):
            print(f"\n❌ Impossible de continuer : le fichier '{output_file}' est ouvert dans Excel.")
            print("Fermez-le puis relancez le programme.")
            exit(1)

        km_rate = float(input("Entrez le taux kilométrique : "))
        meal_price = float(input("Entrez le prix du repas : "))
        loyer = float(input("Entrez le prix du loyer : "))

        wb = Create_weekly_sheets(input_file, year, km_rate, meal_price, loyer)
        Save_workbook_safely(wb , output_file)
    except Exception as e:
        print(f"Une erreur s'est produite : {e}")
