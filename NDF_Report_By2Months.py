# By Arthur Péraud
import os
import openpyxl
from datetime import datetime, timedelta
from openpyxl.workbook import Workbook

def Is_file_locked(filepath: str) -> bool:
    """Vérifie si un fichier est ouvert ou verrouillé (ex: par Excel)."""
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, "a"):
            return False
    except IOError:
        return True

def Get_unique_filename(path : str) -> str:
    """Génère un nom de fichier unique en ajoutant (1), (2), etc. s'il existe déjà."""
    base, ext = os.path.splitext(path)
    i = 1
    new_path = path
    while os.path.exists(new_path):
        new_path = f"{base}({i}){ext}"
        i += 1
    return new_path

def Save_workbook_safely(wb: Workbook, output_file: str) -> None:
    """Sauvegarde fichier Excel, confirmation si existence sinon préfixe est ajouté (Get_unique_filename).
    Crée le dossier si le chemin n'existe pas.
    """

    # Vérifie et crée le dossier si nécessaire
    folder = os.path.dirname(output_file)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
        print(f"📂 Dossier créé : {folder}")

    if os.path.exists(output_file):
        confirm = input(f"\n⚠️  Le fichier '{output_file}' existe déjà. Voulez-vous l’écraser ? (o/n) : ").strip().lower()
        
        if confirm in ['o', 'y']:
            wb.save(output_file)
            print(f"✅ Fichier écrasé et sauvegardé sous {output_file}")
        elif confirm == 'n':
            new_file = Get_unique_filename(output_file)
            wb.save(new_file)
            print(f"📁 Fichier sauvegardé sous un nouveau nom : {new_file}")
        else:
            print("Réponse non reconnue, fichier non sauvegardé.")
    else:
        wb.save(output_file)
        print(f"\n✅ Fichier sauvegardé sous {output_file}")

def Get_number_of_weeks(year : int) -> int:
    """Calcule le nombre de semaines ISO pour une année donnée (52 ou 53 semaines)."""
    # Vérifie le numéro de semaine du dernier jour de l'année
    last_day_of_year = datetime(year, 12, 31)
    last_week = last_day_of_year.isocalendar()[1]

    # S'il y a 53 semaines ISO dans l'année, sinon il y en a 52
    return 53 if last_week == 53 else 52

def Last_week_contains_4_days_of_month(year : int, month : int = 4) -> (bool, int):
    """Retourne un tuple (True/False, numéro de la semaine) si la dernière semaine du mois contient au moins 4 jours du mois."""
    
    # Gérer le cas où le mois est décembre (12)
    if month == 12:
        next_month = 1
        year += 1  # Passage à l'année suivante
    else:
        next_month = month + 1
    
    # Dernier jour du mois
    last_day_of_month = datetime(year, next_month, 1) - timedelta(days=1)
    
    # Le premier jour de la semaine (lundi) pour le dernier jour du mois
    last_week_start = last_day_of_month - timedelta(days=last_day_of_month.weekday())
    
    # Calcul du nombre de jours du mois dans la dernière semaine
    days_in_last_week = 0
    for i in range(7):  # Vérifier les 7 jours de la semaine
        current_day = last_week_start + timedelta(days=i)
        if current_day.month == month:
            days_in_last_week += 1
    
    # Le numéro de la semaine ISO de la dernière semaine du mois
    last_week_number = last_week_start.isocalendar()[1]
    
    # Si la dernière semaine contient 4 jours ou plus du mois
    contains_4_days = days_in_last_week >= 4
    
    return contains_4_days, last_week_number

def Add_brackets_to_filename(path : str) -> str:
    """Rajoute les [] au niveau du nom de fichier"""
    # Trouver la position du dernier backslash (séparateur de dossier)
    last_backslash = path.rfind('\\')
    
    if last_backslash != -1:
        # Extraire la partie avant et après le dernier backslash
        directory = path[:last_backslash + 1]
        filename = path[last_backslash + 1:]
        
        # Ajouter des crochets autour du nom du fichier
        corrected = f"{directory}[{filename}]"
        return corrected
    return f"[{path}]"

def Create_report_sheet(example_file : str, input_file : str, year : int) -> Workbook:
    periods = [
        "Mai/Juin",
        "Juillet/Août",
        "Septembre/Octobre",
        "Novembre/Décembre",
        "Janvier/Février",
        "Mars/Avril"
    ]

    wb = openpyxl.load_workbook(example_file)
    wb2 = openpyxl.load_workbook(input_file)

    # Feuille modèle par défaut
    ws_template = wb.worksheets[0] 
    ws2 = wb2.worksheets[0]  

    sheet_titles = wb2.sheetnames
    start_week = ws2["K1"].value

    start_month = 4 # On commence par le couple Mai/Juin
    nb_weeks = Get_number_of_weeks(year)
    bracket_input_file = Add_brackets_to_filename(input_file)

    # 20XX
    k = 0
    for i in range(6):
        if start_month == 12 : #20XX + 1
            year += 1
            start_month = 0   

        start_month += 2;
        flag, end_week = Last_week_contains_4_days_of_month(year, start_month)

        if not flag : end_week -= 1

        if i not in [4, 5]:
            if end_week == 0: end_week = nb_weeks
            if end_week == 1: end_week = nb_weeks + 1
        else: # 20XX + 1
            end_week += nb_weeks 

        end = end_week - start_week + 1 
        if i == 5 : end = len(sheet_titles) # 20XX + 1

        start_sheet = sheet_titles[k]
        end_sheet = sheet_titles[end - 1]

        formula1 = f"='{bracket_input_file}{start_sheet}'!M30"    
        formula2a = f"='{bracket_input_file}{start_sheet}'!M28"    
        formula2b = f"='{bracket_input_file}{start_sheet}'!M25"    
        formula3 = f"='{bracket_input_file}{start_sheet}'!I30"    
        formula4 = f"='{bracket_input_file}{start_sheet}'!I29"    
        formula5 = f"='{bracket_input_file}{start_sheet}'!M27"  


        # Création des formules de sommes
        for isheet in range(k+1, end):
            sheet = sheet_titles[isheet]

            formula1 += "+" + f"'{bracket_input_file}{sheet}'!M30"    
            formula2a += "+" + f"'{bracket_input_file}{sheet}'!M28"    
            formula2b += "+" + f"'{bracket_input_file}{sheet}'!M25"    
            formula3 += "+" + f"'{bracket_input_file}{sheet}'!I30"    
            formula4 += "+" + f"'{bracket_input_file}{sheet}'!I29"    
            formula5 += "+" + f"'{bracket_input_file}{sheet}'!M27"  
        
        print(formula1, "\n")    

        # Colonne MOIS
        number1 = sheet_titles[k][4] + sheet_titles[k][5] if sheet_titles[k][5] != '_' else sheet_titles[k][4]
        number2 = sheet_titles[end - 1][4] + sheet_titles[end - 1][5] if sheet_titles[end - 1][5] != '_' else sheet_titles[end - 1][4]
        
        ws_template[f"A{27 + i*2}"].value = periods[i] + " (Sem " + number1 + "-" + number2 + ")" 
        ws_template[f"A{5 + i*2}"].value = periods[i] + " (Sem " + number1 + "-" + number2 + ")" 
        
        # Formules Excel
        ws_template[f"B{27 + i*2}"].value = formula1
        ws_template[f"C{27 + i*2}"].value = formula2a
        ws_template[f"C{28 + i*2}"].value = formula2b
        ws_template[f"G{27 + i*2}"].value = formula3
        ws_template[f"H{27 + i*2}"].value = formula4
        ws_template[f"K{27 + i*2}"].value = formula5

        k = end

    # Années 7CV
    ws_template["H21"].value = year - 1   
    ws_template["I21"].value = year   

    print("")
    return wb

### Main Program
if __name__ == "__main__":
    print("Bienvenue dans le Programme Cal Info Mesure du Rapport des Notes de Frais.")
    print("Ecriture des formules : ")
    print("")
    try:
        year = int(input("Entrez l'année au 1er Mai 20XX : "))

        example_file = f"Note de Frais Report_Modele.xlsx" # Fichier Modèle
        input_file = f"E:\\Cal Info Mesure\\Note de Frais\\Année {year }-{year + 1}\\Peraud\\Frais Sem_{year}-{year+1}.xlsx" 
        output_file = f"E:\\Cal Info Mesure\\Invoice&Royalty\\IR{year }-{year + 1}\\Note de Frais Report_{year}-{year+1}.xlsx"
        
        # Arthur
        # example_file = f"C:\\Users\\aznrm\\Desktop\\Programme\\Excel\\Note de Frais Report_2024-2025.xlsx" # Fichier Modèle
        # input_file = f"C:\\Users\\aznrm\\Desktop\\Programme\\Excel\\Frais Sem_{year}-{year+1}.xlsx" 
        # output_file = f"Frais Report_{year}-{year+1}.xlsx"

        # Vérifie si le fichier de sortie est ouvert pour éviter de perdre du temps
        if Is_file_locked(output_file):
            print(f"\n❌ Impossible de continuer : le fichier '{output_file}' est ouvert dans Excel.")
            print("Fermez-le puis relancez le programme.")
            exit(1)
        
        wb = Create_report_sheet(example_file, input_file, year)
        Save_workbook_safely(wb , output_file)
    except Exception as e:
        print(f"Une erreur s'est produite : {e}")



